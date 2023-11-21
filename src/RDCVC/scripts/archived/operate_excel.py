import os

import openpyxl
import xlwings as xw


def integrate_sheet(dir_path: str):
    """整合 excel 数据

    传入目录地址，将目录下的所有 excel 文件整合成一个 excel 文件
    将每个 excel 文件的第一个工作表重命名为 excel 文件名
    然后将所有工作表整合到一个 excel 文件中

    Args:
        dir_path(str): 目录地址
    """
    # 获取目录下所有 excel 文件的文件名 (xlsx,xls)
    excel_names = [
        fname
        for fname in os.listdir(dir_path)
        if fname.endswith(".xlsx") or fname.endswith(".xls")
    ]

    app = xw.App(visible=False, add_book=False)  # 创建一个不可见的 excel 实例
    wb = app.books.add()  # 创建一个新的 excel 文件
    for excel_name in excel_names:
        _wb = xw.Book(os.path.join(dir_path, excel_name))  # 打开 excel 文件
        _ws = _wb.sheets[0]  # 获取 excel 文件的第一个工作表
        _ws.name = excel_name.split("-")[0]  # 重命名 sheet 为 excel 文件名中的第一个 '-' 前的部分
        _ws.api.Copy(After=wb.sheets[-1].api)  # 将源工作表复制到目标工作簿的最后
        _wb.close()  # 关闭源 excel 文件

    wb.save(os.path.join(dir_path, "integrated.xlsx"))  # 保存新的 excel 文件
    wb.close()  # 关闭新的 excel 文件
    app.quit()  # 退出 excel 实例


def consolidate_data(file_path: str, exclude_names: [str], start_row: int = 6):
    """将一个 xls/xlsx 的所有 sheet 的数据整合到一个 sheet 中，并放在第一个 sheet 之前"""
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path)
    consolidated_sheet = wb.create_sheet(title="consolidated", index=0)
    # 遍历其他所有 sheet
    for sheet_name in wb.sheetnames:
        for exclude_name in exclude_names:
            if exclude_name in sheet_name:
                continue

        _sheet = wb[sheet_name]
        # 获取 sheet 的数据（从第 6 行至第一个空行）
        data = []
        for row in _sheet.iter_rows(min_row=start_row):
            if row[0].value is None:
                break
            data.append([cell.value for cell in row])
        # 将数据写入到新建的 sheet 的尾部
        for row_data in data:
            consolidated_sheet.append(row_data)
    # 保存工作簿
    wb.save(file_path)
    wb.close()


def extract_rows_and_clear(
    file_path: str, rows_to_extract: list, new_sheet_suffix: str = "_val"
):
    """
    从 Excel 文件中的每个 sheet 内提取指定的行并保存到新的工作表中。
    同时删除原工作表中已提取的行。

    参数：
        file_path (str): Excel 文件的路径。
        rows_to_extract (list): 要提取的行索引列表。
        new_sheet_suffix (str): 新工作表的名称后缀。

    返回：
        None
    """

    wb = openpyxl.load_workbook(file_path)  # 加载工作簿
    sheet_names = wb.sheetnames

    # 遍历工作簿中的每个工作表
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]  # 获取当前工作表
        extracted_data = []  # 用于存储提取数据的列表
        # 从指定行提取数据
        for row_index in rows_to_extract:
            row_data = []  # 用于存储当前行数据的列表
            for cell in sheet[row_index]:
                row_data.append(cell.value)
            extracted_data.append(row_data)

        # 创建新的工作表并将提取的数据写入其中
        new_sheet_name = sheet_name + new_sheet_suffix
        new_sheet = wb.create_sheet(title=new_sheet_name)
        for row_data in extracted_data:
            new_sheet.append(row_data)

        # 删除已提取的行
        for row_index in reversed(sorted(rows_to_extract)):
            sheet.delete_rows(row_index)

        # 删除空行
        max_row = sheet.max_row
        for row in reversed(range(1, max_row + 1)):
            is_empty_row = all(
                sheet.cell(row=row, column=column).value is None
                for column in range(1, sheet.max_column + 1)
            )
            if is_empty_row:
                sheet.delete_rows(row)

    # 保存修改后的 Excel 文件
    wb.save(file_path)


if __name__ == "__main__":
    
    integrate_sheet(
        r"D:\OneDrive\01 WORK\# DampersClusterControl",
        r"\03 IoT damper fitting\data\IoTDamper_v5\tmp",
    )
    extract_rows_and_clear(
        r"D:\OneDrive\01 WORK\# DampersClusterControl"
        r"\03 IoT damper fitting\data\IoTDamper_v5\tmp\integrated.xlsx",
        [9, 15, 21, 23, 30, 32, 37, 44],
    )
    # consolidate_data(
    #     r'D:\OneDrive\01 WORK\# DampersClusterControl\03 IoT damper '
    #     r'fitting\data\IoTDamper_v5\tmp\integrated.xlsx',
    #     ['consolidated,val'])
